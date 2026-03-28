"""
run_onapp_eval.py — PHA OnApp VL 评测系统

场景：onapp 场景，端侧数据(terminal_data) + 截图 为主要输入，2.0模式（不预注入cloud_data）

5个模型 × 4个条件 × 48个case（有图的）
  条件A: image + terminal_data（基线）
  条件B: A + description.json 页面描述注入
  条件C: A + knowledge.json 健康知识注入
  条件D: A + description + knowledge

每模型一端口，串行执行，session 每 case 清空
裁判: Claude Code 人工（batch_score.py）

用法:
  python run_onapp_eval.py                       # 全量跑（5模型并行 × 4条件串行 × 48 case串行）
  python run_onapp_eval.py --model Qwen3.5-397B  # 只跑指定模型
  python run_onapp_eval.py --condition A         # 只跑指定条件
  python run_onapp_eval.py --retry-failed        # 重跑所有失败/空响应 case
  python run_onapp_eval.py --review              # 查看 checkpoint 状态

前置条件（手动启动5个PHA实例）:
  PHA_NOSTREAM=1 PHA_STATE_DIR=D:/pha-v2/.pha-8010 pha start -p 8010
  PHA_NOSTREAM=1 PHA_STATE_DIR=D:/pha-v2/.pha-8011 pha start -p 8011
  PHA_NOSTREAM=1 PHA_STATE_DIR=D:/pha-v2/.pha-8012 pha start -p 8012
  PHA_NOSTREAM=1 PHA_STATE_DIR=D:/pha-v2/.pha-8013 pha start -p 8013
  PHA_NOSTREAM=1 PHA_STATE_DIR=D:/pha-v2/.pha-8014 pha start -p 8014
"""

import os, sys, base64, json, time, uuid, threading, shutil, argparse
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import openpyxl
from PIL import Image
import io

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if sys.stderr.encoding != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

os.environ["NO_PROXY"] = "127.0.0.1,localhost"
os.environ["no_proxy"] = "127.0.0.1,localhost"
os.environ["HTTP_PROXY"]  = ""
os.environ["HTTPS_PROXY"] = ""
os.environ["http_proxy"]  = ""
os.environ["https_proxy"] = ""
NO_PROXY = {"http": None, "https": None}

# ── Paths ─────────────────────────────────────────────────────────────────────
EVAL_DIR      = Path(__file__).parent
PHA_ROOT      = Path(r"D:\pha-v2")
PICS_DIR      = EVAL_DIR / "pics"
XLSX_PATH     = EVAL_DIR / "demotest.xlsx"
DESC_PATH     = EVAL_DIR / "description.json"
KW_DICT_PATH  = EVAL_DIR / "keyword_dict.json.json"
KNOWLEDGE_PATH= EVAL_DIR / "knowledge.json"

RESULTS_BASE  = EVAL_DIR / "results"
RUN_ID        = time.strftime("run_%Y%m%d_%H%M")
RESULTS_DIR   = RESULTS_BASE / RUN_ID
FAILED_LOG    = RESULTS_DIR / "FAILED_CASES.json"

OPENROUTER_KEY = "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"

# ── Models ────────────────────────────────────────────────────────────────────
MODELS = [
    {"id": "qwen/qwen3.5-397b-a17b",          "name": "Qwen3.5-397B", "port": 8010},
    {"id": "qwen/qwen3.5-122b-a10b",           "name": "Qwen3.5-122B", "port": 8011},
    {"id": "qwen/qwen3-vl-235b-a22b-instruct", "name": "Qwen3-235B",   "port": 8012},
    {"id": "moonshotai/kimi-k2.5",             "name": "kimi-k2.5",    "port": 8013},
    {"id": "z-ai/glm-4.6v",                    "name": "GLM-4.6V",     "port": 8014},
]

CONDITIONS = ["A", "B", "C", "D"]

# 12个评测维度（初始化用，实际评分由batch_score.py填入）
DIMS = [
    "视觉识别准确率", "幻觉控制率", "数值读取精度", "输出时序合规",
    "安全声明合规",   "边界克制合规", "数据引用质量", "端侧数据优先性",
    "工具调用时机准确性", "工具调用结果整合度", "任务完成度", "图像与上下文一致性",
]

IMG_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
PRINT_LOCK = threading.Lock()


# ── Static data loading ───────────────────────────────────────────────────────
def load_static_data():
    desc = json.loads(DESC_PATH.read_text("utf-8"))
    kw_dict = json.loads(KW_DICT_PATH.read_text("utf-8"))
    knowledge = json.loads(KNOWLEDGE_PATH.read_text("utf-8"))
    return desc, kw_dict, knowledge


def get_description(current_page: str, desc: dict) -> str:
    return desc.get(current_page, "")


def get_knowledge(current_page: str, kw_dict: dict, knowledge: dict) -> str:
    keywords = kw_dict.get(current_page, [])
    texts = []
    for kw in keywords:
        if kw in knowledge:
            texts.append(f"【{kw}】\n{knowledge[kw]}")
    return "\n\n".join(texts)


# ── Case loading ──────────────────────────────────────────────────────────────
def load_cases() -> list:
    """Load all test cases from demotest.xlsx Sheet1 where image is available."""
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb["Sheet1"]
    available_imgs = set(os.listdir(str(PICS_DIR)))

    cases = []
    for r in range(2, ws.max_row + 1):
        query      = ws.cell(r, 1).value
        image_name = ws.cell(r, 2).value
        td_raw     = ws.cell(r, 3).value
        cloud_data = ws.cell(r, 4).value

        if not image_name or image_name not in available_imgs:
            continue
        if not query or not td_raw:
            continue

        try:
            td_obj = json.loads(td_raw)
        except Exception:
            td_obj = {}

        current_page = td_obj.get("currentPage", "")
        case_id = f"row{r:03d}"

        cases.append({
            "case_id":      case_id,
            "row":          r,
            "query":        query,
            "image_name":   image_name,
            "current_page": current_page,
            "terminal_data": td_raw,
            "terminal_data_obj": td_obj,
            "cloud_data":   cloud_data or "",
        })

    log(f"[cases] loaded {len(cases)} cases with available images")
    return cases


# ── PHA per-port setup ────────────────────────────────────────────────────────
def setup_pha_port(model_cfg: dict):
    """Write config.json for this port's .pha directory."""
    port = model_cfg["port"]
    pha_dir = PHA_ROOT / f".pha-{port}"
    pha_dir.mkdir(parents=True, exist_ok=True)
    (pha_dir / "users" / "anonymous" / "sessions").mkdir(parents=True, exist_ok=True)

    config = {
        "gateway": {"host": "0.0.0.0", "port": port, "autoStart": False},
        "dataSources": {"type": "mock"},
        "tui": {"theme": "dark", "showToolCalls": True},
        "orchestrator": {"pha": "openrouter/eval-model"},
        "models": {
            "providers": {
                "openrouter": {
                    "models": [{"name": "eval-model", "model": model_cfg["id"]}],
                    "apiKey": OPENROUTER_KEY,
                    "baseUrl": "https://openrouter.ai/api/v1",
                }
            }
        },
    }
    cfg_path = pha_dir / "config.json"
    cfg_path.write_text(json.dumps(config, indent=2, ensure_ascii=False), "utf-8")
    log(f"[setup] port {port} ({model_cfg['name']}) → {cfg_path}")


def check_pha_health(port: int, timeout: int = 10) -> bool:
    try:
        r = requests.get(f"http://127.0.0.1:{port}/health",
                         proxies=NO_PROXY, timeout=timeout)
        return r.status_code == 200
    except Exception:
        return False


def clear_sessions(port: int):
    """Clear session files to prevent context bleed between cases."""
    sessions_dir = PHA_ROOT / f".pha-{port}" / "users" / "anonymous" / "sessions"
    if sessions_dir.exists():
        shutil.rmtree(str(sessions_dir))
    sessions_dir.mkdir(parents=True, exist_ok=True)


# ── Image helpers ─────────────────────────────────────────────────────────────
def encode_image(path: Path) -> tuple:
    """Returns (base64_str, mime_type)."""
    suffix = path.suffix.lower()
    mime_map = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
        ".webp": "image/webp", ".gif": "image/gif", ".bmp": "image/bmp",
    }
    mime = mime_map.get(suffix, "image/jpeg")
    data = path.read_bytes()
    # Resize if > 3MB
    if len(data) > 3 * 1024 * 1024:
        img = Image.open(io.BytesIO(data)).convert("RGB")
        w, h = img.size
        scale = 1200 / max(w, h)
        if scale < 1:
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        data, mime = buf.getvalue(), "image/jpeg"
    return base64.b64encode(data).decode(), mime


# ── PHA API ───────────────────────────────────────────────────────────────────
def upload_image(port: int, b64: str, mime: str, retries: int = 3) -> str | None:
    for attempt in range(retries):
        try:
            r = requests.post(
                f"http://127.0.0.1:{port}/api/upload/diet-photo",
                json={"imageBase64": b64, "mimeType": mime},
                proxies=NO_PROXY, timeout=30,
            )
            iid = r.json().get("imageId")
            if iid:
                return iid
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
            else:
                log(f"  [upload failed] port={port} {e}")
    return None


def call_pha(port: int, content: str, image_id: str | None,
             retries: int = 3) -> tuple[str, list]:
    """Returns (response_text, tool_calls_list)."""
    if image_id:
        full_content = f"[vision] image_ids={image_id} {content}"
    else:
        full_content = content

    payload = {
        "messages": [{"role": "user", "content": full_content}],
        "thread_id": str(uuid.uuid4()),
        "run_id": str(uuid.uuid4()),
        "context": [],
    }

    for attempt in range(retries):
        text_parts = []
        tool_calls = []
        current_tool = None

        try:
            with requests.post(
                f"http://127.0.0.1:{port}/api/ag-ui",
                json=payload,
                headers={"Accept": "text/event-stream", "Content-Type": "application/json"},
                proxies=NO_PROXY, stream=True, timeout=180,
            ) as resp:
                for raw in resp.iter_lines():
                    line = raw.decode("utf-8", "replace") if isinstance(raw, bytes) else raw
                    if not line or not line.startswith("data:"):
                        continue
                    ds = line[5:].strip()
                    if ds == "[DONE]":
                        break
                    try:
                        evt = json.loads(ds)
                        t = evt.get("type", "")
                        if t == "TextMessageContent":
                            text_parts.append(evt.get("delta", ""))
                        elif t == "ToolCallStart":
                            current_tool = {
                                "name": evt.get("toolCallName", evt.get("name", "unknown")),
                                "args": "",
                                "result": "",
                            }
                        elif t == "ToolCallArgsChunk" and current_tool is not None:
                            current_tool["args"] += evt.get("delta", "")
                        elif t == "ToolCallEnd" and current_tool is not None:
                            # Try parse args as JSON
                            try:
                                current_tool["args"] = json.loads(current_tool["args"])
                            except Exception:
                                pass
                            tool_calls.append(current_tool)
                            current_tool = None
                        elif t == "RunFinished":
                            break
                        elif t == "RunError":
                            log(f"  [RunError] port={port}: {evt.get('message','')}")
                            break
                    except Exception:
                        pass

            text = "".join(text_parts).strip()
            if text:
                return text, tool_calls
            if attempt < retries - 1:
                time.sleep(3 * (attempt + 1))

        except Exception as e:
            if attempt < retries - 1:
                time.sleep(3 * (attempt + 1))
            else:
                log(f"  [chat error] port={port} after {retries} tries: {e}")

    return "".join(text_parts).strip(), tool_calls


# ── Prompt construction ───────────────────────────────────────────────────────
def build_prompt(case: dict, condition: str, desc: dict, kw_dict: dict,
                 knowledge: dict) -> tuple[str, dict]:
    """Returns (prompt_content, injected_dict)."""
    query        = case["query"]
    td_raw       = case["terminal_data"]
    current_page = case["current_page"]

    injected = {"description": None, "knowledge": None}

    parts = [query, "\n\n【当前页面数据】\n" + td_raw]

    if condition in ("B", "D"):
        d = get_description(current_page, desc)
        if d:
            parts.append("\n\n【页面界面说明】\n" + d)
            injected["description"] = d[:200] + "..." if len(d) > 200 else d

    if condition in ("C", "D"):
        k = get_knowledge(current_page, kw_dict, knowledge)
        if k:
            parts.append("\n\n【健康知识参考】\n" + k)
            injected["knowledge"] = k[:200] + "..." if len(k) > 200 else k

    return "".join(parts), injected


# ── Checkpoint ────────────────────────────────────────────────────────────────
def load_checkpoint(json_path: Path) -> dict:
    """Returns {case_id: result_dict} per condition."""
    if not json_path.exists():
        return {}
    try:
        data = json.loads(json_path.read_text("utf-8"))
        idx = {}
        for cond, results in data.get("conditions", {}).items():
            for r in results:
                key = f"{cond}::{r['case_id']}"
                idx[key] = r
        return idx
    except Exception:
        return {}


def save_checkpoint(json_path: Path, model_cfg: dict, cond_results: dict):
    """cond_results: {"A": [result,...], "B": [...], ...}"""
    data = {
        "model": model_cfg,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "conditions": cond_results,
    }
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), "utf-8")


def append_failed(model_name: str, condition: str, case_id: str, error: str):
    failed = {}
    if FAILED_LOG.exists():
        try:
            failed = json.loads(FAILED_LOG.read_text("utf-8"))
        except Exception:
            pass
    key = f"{model_name}::{condition}::{case_id}"
    failed[key] = {
        "model": model_name, "condition": condition, "case_id": case_id,
        "error": error, "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    FAILED_LOG.write_text(json.dumps(failed, ensure_ascii=False, indent=2), "utf-8")


def is_zombie(result: dict) -> bool:
    if result.get("status") == "failed":
        return True
    if not result.get("response", ""):
        return True
    # Token loop: response > 5000 chars with repetition
    resp = result.get("response", "")
    if len(resp) > 5000:
        chunk = resp[:100]
        if resp.count(chunk) > 3:
            return True
    return False


# ── Logging ───────────────────────────────────────────────────────────────────
def log(msg: str):
    with PRINT_LOCK:
        ts = time.strftime("%H:%M:%S")
        print(f"[{ts}] {msg}", flush=True)


# ── Single model runner ───────────────────────────────────────────────────────
def run_model(model_cfg: dict, all_cases: list, target_conditions: list,
              force_retry: bool, desc: dict, kw_dict: dict, knowledge: dict):
    name  = model_cfg["name"]
    port  = model_cfg["port"]
    json_path = RESULTS_DIR / f"{name}_results.json"

    # Check PHA health
    if not check_pha_health(port):
        log(f"[{name}] ⚠️  Port {port} not healthy — SKIP. Start PHA first.")
        return

    log(f"[{name}] Port {port} healthy ✓")

    # Load checkpoint
    checkpoint = load_checkpoint(json_path)
    cond_results: dict[str, list] = {c: [] for c in CONDITIONS}
    # Restore existing condition results from checkpoint
    if json_path.exists():
        try:
            existing = json.loads(json_path.read_text("utf-8"))
            for cond, results in existing.get("conditions", {}).items():
                cond_results[cond] = results
        except Exception:
            pass

    for condition in target_conditions:
        log(f"[{name}][{condition}] Starting {len(all_cases)} cases...")

        # Build index of existing results for this condition
        done_map: dict[str, dict] = {}
        for r in cond_results.get(condition, []):
            done_map[r["case_id"]] = r

        # Determine which cases to run
        if force_retry:
            to_run = all_cases
        else:
            to_run = [c for c in all_cases
                      if c["case_id"] not in done_map or is_zombie(done_map[c["case_id"]])]

        skipped = len(all_cases) - len(to_run)
        if skipped:
            log(f"[{name}][{condition}] Skipping {skipped} cached, running {len(to_run)}")

        new_results: list[dict] = []

        for i, case in enumerate(to_run):
            case_id    = case["case_id"]
            image_name = case["image_name"]
            short_q    = case["query"][:20]

            log(f"[{name}][{condition}][{i+1}/{len(to_run)}] {case_id} | {image_name} | {short_q}")

            # Clear sessions
            clear_sessions(port)

            # Upload image
            img_path = PICS_DIR / image_name
            b64, mime = encode_image(img_path)
            image_id = upload_image(port, b64, mime)
            if not image_id:
                log(f"  [{name}][{condition}][{case_id}] ⚠️  Upload failed")
                append_failed(name, condition, case_id, "image upload failed")
                result = _failed_result(case, condition, "image upload failed")
                new_results.append(result)
                done_map[case_id] = result
                _flush_cond(cond_results, condition, done_map, all_cases)
                save_checkpoint(json_path, model_cfg, cond_results)
                continue

            # Build prompt
            prompt, injected = build_prompt(case, condition, desc, kw_dict, knowledge)

            # Call PHA
            response, tool_calls = call_pha(port, prompt, image_id)

            if not response:
                log(f"  [{name}][{condition}][{case_id}] ⚠️  Empty response")
                append_failed(name, condition, case_id, "empty response")
                result = _failed_result(case, condition, "empty response")
            else:
                # Detect token loop
                status = "done"
                if len(response) > 5000 and response.count(response[:100]) > 3:
                    log(f"  [{name}][{condition}][{case_id}] ⚠️  Token loop detected ({len(response)}c)")
                    status = "token_loop"

                log(f"  [{name}][{condition}][{case_id}] ✓ {len(response)}c | tools={len(tool_calls)}: {[t['name'] for t in tool_calls]}")

                result = {
                    "case_id":      case_id,
                    "row":          case["row"],
                    "query":        case["query"],
                    "image_name":   image_name,
                    "current_page": case["current_page"],
                    "terminal_data": case["terminal_data"],
                    "cloud_data":   case["cloud_data"],
                    "condition":    condition,
                    "injected":     injected,
                    "response":     response,
                    "tool_calls":   tool_calls,
                    "status":       status,
                    "judgment":     None,  # filled by batch_score.py
                }

            new_results.append(result)
            done_map[case_id] = result
            _flush_cond(cond_results, condition, done_map, all_cases)
            save_checkpoint(json_path, model_cfg, cond_results)
            time.sleep(0.5)  # brief pause between cases

        log(f"[{name}][{condition}] Done ✓ ({len(new_results)} new results)")

    log(f"[{name}] All conditions complete → {json_path}")


def _flush_cond(cond_results: dict, condition: str, done_map: dict, all_cases: list):
    """Rebuild ordered list for this condition from done_map."""
    cond_results[condition] = [done_map[c["case_id"]] for c in all_cases
                               if c["case_id"] in done_map]


def _failed_result(case: dict, condition: str, error: str) -> dict:
    return {
        "case_id":      case["case_id"],
        "row":          case["row"],
        "query":        case["query"],
        "image_name":   case["image_name"],
        "current_page": case["current_page"],
        "terminal_data": case["terminal_data"],
        "cloud_data":   case["cloud_data"],
        "condition":    condition,
        "injected":     {},
        "response":     "",
        "tool_calls":   [],
        "status":       "failed",
        "error":        error,
        "judgment":     None,
    }


# ── Review ────────────────────────────────────────────────────────────────────
def review_all():
    if not RESULTS_DIR.exists():
        # Find latest run
        runs = sorted(RESULTS_BASE.glob("run_*"), reverse=True)
        if not runs:
            print("No results found.")
            return
        target = runs[0]
    else:
        target = RESULTS_DIR

    print(f"\n=== Checkpoint review: {target} ===\n")
    for json_path in sorted(target.glob("*_results.json")):
        data = json.loads(json_path.read_text("utf-8"))
        name = data["model"]["name"]
        print(f"  {name}:")
        for cond, results in data.get("conditions", {}).items():
            done    = sum(1 for r in results if r.get("status") == "done")
            failed  = sum(1 for r in results if r.get("status") == "failed")
            loops   = sum(1 for r in results if r.get("status") == "token_loop")
            judged  = sum(1 for r in results if r.get("judgment") is not None)
            zombies = [r["case_id"] for r in results if is_zombie(r)]
            print(f"    [{cond}] done={done} failed={failed} loops={loops} "
                  f"judged={judged} zombies={len(zombies)}")
            if zombies:
                print(f"         zombies: {', '.join(zombies)}")
    print()


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", help="Only run this model name")
    parser.add_argument("--condition", help="Only run this condition (A/B/C/D)")
    parser.add_argument("--retry-failed", action="store_true")
    parser.add_argument("--review", action="store_true")
    args = parser.parse_args()

    if args.review:
        review_all()
        return

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    # Load static data
    desc, kw_dict, knowledge = load_static_data()
    log(f"[init] description={len(desc)} pages, knowledge={len(knowledge)} topics")

    # Load cases
    cases = load_cases()
    if not cases:
        log("[init] ⚠️  No cases found — check pics/ directory")
        return

    # Filter models
    models = MODELS
    if args.model:
        models = [m for m in MODELS if m["name"] == args.model]
        if not models:
            log(f"[init] Unknown model: {args.model}")
            return

    # Filter conditions
    target_conditions = CONDITIONS
    if args.condition:
        if args.condition not in CONDITIONS:
            log(f"[init] Unknown condition: {args.condition}")
            return
        target_conditions = [args.condition]

    # Setup .pha dirs
    for m in models:
        setup_pha_port(m)

    log(f"[init] Run ID: {RUN_ID}")
    log(f"[init] Models: {[m['name'] for m in models]}")
    log(f"[init] Conditions: {target_conditions}")
    log(f"[init] Cases: {len(cases)}")
    log(f"[init] Total runs: {len(models) * len(target_conditions) * len(cases)}")
    log(f"[init] Results dir: {RESULTS_DIR}")
    log("")

    # Run models in parallel (one thread per model/port)
    with ThreadPoolExecutor(max_workers=len(models)) as executor:
        futures = {
            executor.submit(
                run_model, m, cases, target_conditions,
                args.retry_failed, desc, kw_dict, knowledge
            ): m["name"]
            for m in models
        }
        for future in as_completed(futures):
            name = futures[future]
            try:
                future.result()
            except Exception as e:
                log(f"[{name}] ❌ Thread error: {e}")

    log(f"\n[done] All results in {RESULTS_DIR}")
    log("[done] Next step: python batch_score.py to score responses")
    review_all()


if __name__ == "__main__":
    main()
